import pandas as pd
from openpyxl import load_workbook


def unit_totals(path, dest):
    unit_report = pd.read_excel(path)
    pd.set_option('display.width', 400)
    pd.set_option('display.max_columns', None)
    unit_report.columns = ['Mask', 'Date', 'Unit', 'Qty']
    unit_report = unit_report[['Unit', 'Date', 'Mask', 'Qty']]
    unit_report['Mask'] = unit_report['Mask'].astype('category')
    unit_report['Date'] = unit_report['Date'].dt.date

    unit_report['Mask'].replace({
        'FILTER PFR95 RESPIRATOR REG SIZE': 'Halyard Duckbill Reg',
        'GERSON N95 MASK RESPIRATOR 1730': 'Gerson 1730',
        'GERSON N95 MASK RESPIRATOR 82130': 'Gerson 82130',
        'MASK FACE RESPIRATOR N95 SMALL BX/20EA': '3M 1860S Small',
        'MASK RESPIRATOR PARTICULATE CONE N95 NIOSH CERTIFIED FLUID R': '3M 1860 Regular',
        'MASK RESPIRATOR SM CS/6BX/35EA': 'Halyard Duckbill Small'
    }, inplace=True)

    unit_report = unit_report.set_index("Date")
    unit_report.to_excel(path)
    writer = pd.ExcelWriter(path, engine="openpyxl")
    book = load_workbook(path)
    writer.book = book
    sheet = book.get_sheet_by_name("Sheet1")
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 14
    writer.save()
    writer.close()

    # Unit Totals
    unit_total = unit_report[['Unit', 'Qty']]
    unit_total.columns = ['Unit', 'Unit Total']
    units = unit_total.groupby('Unit')
    df = pd.DataFrame(units.sum()).sort_values('Unit Total', ascending=False)

    # Add Units Totals Sheet to N95 Report
    writer = pd.ExcelWriter(dest, engine="openpyxl")
    book = load_workbook(dest)
    writer.book = book
    df.to_excel(writer, sheet_name="Unit Totals")
    sheet = book.get_sheet_by_name("Unit Totals")
    sheet.column_dimensions["A"].width = 30
    writer.save()
    writer.close()
