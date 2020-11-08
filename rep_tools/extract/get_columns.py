from openpyxl import load_workbook
import pandas as pd
import os

desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), r'OneDrive\Desktop')
src_file = r"{}".format(desktop) + r"\shipping_tables.xlsx"

# Loads the workbook
wb = load_workbook(filename=src_file)

# See all the sheets within the workbook
# print(wb.sheetnames)

# See a list of all named tables
# sheet = wb['shipping_rates']
# print(sheet.tables.keys())

# Lookup the range of a "Name Excel Table"
sheet = wb['shipping_rates']
lookup_table = sheet.tables['ship_cost']
print(lookup_table.ref)

# Access the data in the table range
data = sheet[lookup_table.ref]
print(data)

# Convert tuple into a DataFrame
rows_list = []

# Loop through each row and get the values in the cells
for row in data:
    # Get a list of all columns in each row
    cols = []
    for col in row:
        cols.append(col.value)
    rows_list.append(cols)

# Create a pandas dataframe from the rows_list.
# The first row is the column names
df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])
